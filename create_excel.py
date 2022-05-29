import openpyxl as px

#PatternFillで背景を塗りつぶす、Borderで罫線の描画、Sideで罫線のスタイルの指定、Fontでフォント指定
from openpyxl.styles import PatternFill,Border,Side,Font

#新規作成(オブジェクト生成)から保存。
wb = px.Workbook()
wb.save('test.xlsx')

#ファイルを読み込み
wb          = px.load_workbook('test.xlsx')

#アクティブシートを選択(新規作成時に最初からあるシート)
ws          = wb.active

#シート名を変更
ws.title    = "領収書"

#セルに値を入力していく
import datetime

ws["A1"].value  = "決済日"
ws["A2"].value  = datetime.date.today()

ws["B1"].value  = "商品名"
ws["C1"].value  = "個数"
ws["D1"].value  = "小計"

ws["B2"].value  = "商品A"
ws["C2"].value  = 2
ws["D2"].value  = 20000

ws["B3"].value  = "商品B"
ws["C3"].value  = 1
ws["D3"].value  = 30000


#セルの値を入手。計算結果を入力
ws["F2"].value  = "請求金額"
ws["F3"].value  = "=SUM(C2*D2,C3*D3)"


rows    = ws.iter_rows()
print(rows)

#行ごとにループさせる。
for row in rows:
    print(row)
    for cell in row:
        print(cell.value)
        
        #参照元:https://openpyxl.readthedocs.io/en/stable/styles.html
        
        #背景色を塗りつぶす
        cell.fill   = PatternFill(fgColor="FFFF00",fill_type="solid")

        #Sideで予め罫線のオブジェクトを作っておく。
        thin        = Side(border_style="thin", color="000000")

        #Borderで罫線を描画
        #cell.border = Border(left=thin,right=thin,top=thin,bottom=thin)

        #↓と↑は等価

        cell.border = Border(thin,thin,thin,thin)
        
        #Fontでフォントの装飾、b=Trueで太字化。sizeでサイズ指定
        cell.font   = Font(b=True,size=15)



#特定文字列を含むセルを検索
rows    = ws.iter_rows()

for row in rows:
    for cell in row:
        if cell.value == "請求金額":
            cell.value = "お支払金額"

    
wb.save('test.xlsx')

